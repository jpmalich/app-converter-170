"""ISS supplier-admin pricing endpoints.

Mirrors `pricing_admin.py` but for the single-tier ISS catalog (collection
`iss_catalog`). Workflow:

  1. GET  /admin/iss/export   → download current prices as CSV
  2. POST /admin/iss/upload   → upload CSV/XLSX, returns a diff (no save)
  3. POST /admin/iss/apply    → commit the diff returned by /upload

CSV columns: section, name, unit, price

The supplier may also edit the file in Excel and re-upload — only rows
whose `price` cell changed appear in the diff, so untouched rows are
silently ignored.

ISS prices float by line, not by a flat %, so there is no "% bump"
endpoint here — uploads are the sole admin-side update path.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from db import db
from deps import check_admin_token
from iss_catalog import ensure_iss_catalog_seeded

router = APIRouter()

CSV_COLUMNS = ["section", "name", "unit", "price"]


class ISSPriceChange(BaseModel):
    """Single proposed price change. Frontend holds the diff returned by
    /upload in component state and POSTs it back to /apply unchanged."""
    section: str
    name: str
    unit: str
    old: float
    new: float


class ISSApplyIn(BaseModel):
    changes: list[ISSPriceChange]


def _round2(x: float) -> float:
    return round(float(x) + 0.0, 2)


def _parse_upload(filename: str, raw: bytes) -> list[dict]:
    """Returns normalized rows: list of {section, name, unit, price}."""
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    rows: list[dict] = []

    if ext in {"xlsx", "xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise HTTPException(status_code=500, detail="openpyxl not installed") from e
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
        header_row: Optional[list[str]] = None
        for row in ws.iter_rows(values_only=True):
            if header_row is None:
                header_row = [str(c or "").strip().lower() for c in row]
                missing = [c for c in CSV_COLUMNS if c not in header_row]
                if missing:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Spreadsheet missing required columns: {', '.join(missing)}",
                    )
                continue
            d = {header_row[i]: row[i] for i in range(min(len(header_row), len(row)))}
            if not any(d.get(c) for c in ("section", "name")):
                continue
            rows.append(d)
    else:
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise HTTPException(status_code=400, detail="Empty CSV (no header row)")
        normalized = [(f or "").strip().lower() for f in reader.fieldnames]
        missing = [c for c in CSV_COLUMNS if c not in normalized]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"CSV missing required columns: {', '.join(missing)}",
            )
        text2 = ",".join(normalized) + "\n" + text.split("\n", 1)[1]
        reader = csv.DictReader(io.StringIO(text2))
        for r in reader:
            rows.append(r)
    return rows


async def _diff_upload(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Compute (changes, unmatched). Unmatched rows are surfaced in the
    UI so the supplier can fix typos before applying."""
    cursor = db.iss_catalog.find({}, {"_id": 0})
    docs = await cursor.to_list(500)
    lookup: dict[tuple[str, str], dict] = {
        (d["section"].strip().lower(), d["name"].strip().lower()): d
        for d in docs
    }

    changes: list[dict] = []
    unmatched: list[dict] = []
    for i, r in enumerate(rows, start=2):  # row 1 was the header
        section_v = (r.get("section") or "").strip()
        name_v = (r.get("name") or "").strip()
        if not (section_v and name_v):
            unmatched.append({"row": i, "reason": "Missing section/name"})
            continue
        key = (section_v.lower(), name_v.lower())
        found = lookup.get(key)
        if not found:
            unmatched.append({
                "row": i, "section": section_v, "name": name_v,
                "reason": "No matching catalog item — check spelling",
            })
            continue
        raw_val = r.get("price")
        if raw_val in (None, ""):
            continue
        try:
            new = _round2(float(str(raw_val).replace("$", "").replace(",", "")))
        except (TypeError, ValueError):
            unmatched.append({
                "row": i, "section": section_v, "name": name_v,
                "reason": f"Non-numeric price: {raw_val!r}",
            })
            continue
        old = float(found.get("price", 0) or 0)
        if abs(new - old) < 0.005:
            continue
        changes.append({
            "section": found["section"],
            "name": found["name"],
            "unit": found.get("unit", ""),
            "old": old,
            "new": new,
        })
    return changes, unmatched


async def _apply_changes(changes: list[dict]) -> int:
    applied = 0
    now_iso = datetime.now(timezone.utc).isoformat()
    for c in changes:
        res = await db.iss_catalog.update_one(
            {"section": c["section"], "name": c["name"]},
            {"$set": {
                "price": _round2(c["new"]),
                "updated_at": now_iso,
            }},
        )
        if res.modified_count:
            applied += 1
    return applied


@router.get("/admin/iss/export")
async def export_iss_pricing(request: Request):
    """Download the current ISS catalog as CSV. Edit in Excel and re-upload."""
    check_admin_token(request)
    await ensure_iss_catalog_seeded(db)
    cursor = db.iss_catalog.find({}, {"_id": 0}).sort([
        ("section_order", 1),
        ("item_order", 1),
    ])
    docs = await cursor.to_list(500)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(CSV_COLUMNS)
    for d in docs:
        writer.writerow([
            d["section"], d["name"], d.get("unit", ""),
            f"{float(d.get('price', 0) or 0):.2f}",
        ])
    buf.seek(0)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="iss-pricing-{today}.csv"'},
    )


@router.post("/admin/iss/upload")
async def upload_iss_pricing(
    request: Request,
    file: UploadFile = File(...),
    commit: str = Form("false"),
):
    """Preview an uploaded CSV/XLSX. Pass `commit=true` to apply immediately."""
    check_admin_token(request)
    await ensure_iss_catalog_seeded(db)
    raw = await file.read()
    if len(raw) > 4 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (>4MB)")
    rows = _parse_upload(file.filename or "", raw)
    changes, unmatched = await _diff_upload(rows)
    applied = 0
    if commit.lower() == "true":
        applied = await _apply_changes(changes)
    return {"changes": changes, "unmatched": unmatched, "applied": applied}


@router.post("/admin/iss/apply")
async def apply_iss_changes(body: ISSApplyIn, request: Request):
    """Commit a previewed changeset."""
    check_admin_token(request)
    if not body.changes:
        return {"applied": 0}
    applied = await _apply_changes([c.model_dump() for c in body.changes])
    return {"applied": applied}
